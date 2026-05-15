"""
score_sequences.py

Computes a comprehensive developability feature panel for all generated sequences.

Features computed:
  Sequence-level:
    - Physicochemical: pI, net charge, MW, GRAVY, aromaticity
    - Humanness: OASis score, germline identity %, Sapiens per-position scores
    - CDR properties: lengths, charge, hydrophobicity, composition
    - Sequence liabilities: deamidation, oxidation, isomerization,
                            glycosylation, unpaired Cys, Asp-Pro
                            (each split into CDR count, FR count, and
                             per-flag detail with IMGT position and region)
  Vernier zone:
    - Identity vs mouse, vs lab Hu, vs lab final
    - Back-mutation count, humanized count
    - Per-position mismatch list

  Structure-level (optional, requires IgFold):
    - pLDDT per region (CDR1/2/3, FR1/2/3/4)
    - VH-VL interface contacts

Usage:
    python3 evaluation/score_sequences.py \\
        --generated outputs/all_sequences.csv \\
        --benchmark data/benchmarks/humanization_benchmark.csv \\
        --output outputs/scores.csv

    # With structure prediction:
    python3 evaluation/score_sequences.py \\
        --generated outputs/all_sequences.csv \\
        --benchmark data/benchmarks/humanization_benchmark.csv \\
        --output outputs/scores.csv \\
        --structure
"""

# isort: skip_file
import sys
sys.path.insert(0, "/workspace/antibody-humanization-tool")  # noqa: E402
from pipeline.step_a_numbering import number_sequence
from typing import Optional
import argparse
import csv
import re
import math


# ── Vernier zone positions (IMGT numbering) ───────────────────────────────────
# These FR positions physically support CDR loop conformation.
# Source: Foote & Winter (1992), Chothia et al. (1998)

VERNIER_VH = {2, 27, 29, 30, 47, 48, 67, 69, 71, 78, 80, 93, 94}
VERNIER_VL = {2, 4, 35, 36, 46, 47, 48, 49, 64, 66, 68, 69, 71}


# ── Amino acid property tables ────────────────────────────────────────────────

AA_CHARGE_PH7 = {
    'R': +1, 'K': +1, 'H': +0.1,
    'D': -1, 'E': -1,
}

AA_HYDROPHOBICITY = {  # Kyte-Doolittle scale
    'A':  1.8, 'R': -4.5, 'N': -3.5, 'D': -3.5, 'C':  2.5,
    'Q': -3.5, 'E': -3.5, 'G': -0.4, 'H': -3.2, 'I':  4.5,
    'L':  3.8, 'K': -3.9, 'M':  1.9, 'F':  2.8, 'P': -1.6,
    'S': -0.8, 'T': -0.7, 'W': -0.9, 'Y': -1.3, 'V':  4.2,
}

AA_MW = {  # monoisotopic residue masses
    'A': 71.04, 'R': 156.10, 'N': 114.04, 'D': 115.03, 'C': 103.01,
    'Q': 128.06, 'E': 129.04, 'G': 57.02,  'H': 137.06, 'I': 113.08,
    'L': 113.08, 'K': 128.09, 'M': 131.04, 'F': 147.07, 'P': 97.05,
    'S': 87.03,  'T': 101.05, 'W': 186.08, 'Y': 163.06, 'V': 99.07,
}


# ── Sequence liability detection ──────────────────────────────────────────────

DEAMIDATION_MOTIFS = re.compile(r'N[GSTCA]')  # NG, NS, NT, NC, NA
OXIDATION_MOTIFS = re.compile(r'[MW]')
ISOMERIZATION_MOTIFS = re.compile(r'D[GSDA]')   # DG, DS, DD, DA
GLYCOSYLATION_MOTIF = re.compile(r'N[^P][ST]')  # N-X-S/T, X≠P
ASP_PRO_MOTIF = re.compile(r'DP')

# Deamidation motif risk ranking — NG is highest risk, others are moderate
DEAMIDATION_RISK = {'NG': 'high', 'NS': 'moderate', 'NT': 'moderate',
                    'NC': 'moderate', 'NA': 'moderate'}


def _str_pos_to_imgt(str_pos: int, numbered: dict) -> Optional[tuple]:
    """
    Map a string position (0-indexed) back to its IMGT position and region.

    The numbered sequence from number_sequence() stores residues as:
      fr_residues:  {imgt_pos: aa, ...}
      cdr_residues: {imgt_pos: aa, ...}  (insertions as tuples e.g. (111,'A'))

    We reconstruct the ordered position list and map by index.

    Returns (imgt_pos, region) or None if mapping fails.
    """
    try:
        def _sk(pos):
            """Sort key for IMGT positions including insertion tuples."""
            if isinstance(pos, tuple):
                return (pos[0], pos[1])
            return (pos, ' ')

        # Build ordered list of (imgt_pos, region) in sequence order
        ordered = []
        for pos, aa in numbered['fr_residues'].items():
            ordered.append((pos, 'FR'))
        for pos, aa in numbered['cdr_residues'].items():
            ordered.append((pos, 'CDR'))
        ordered.sort(key=lambda x: _sk(x[0]))

        if str_pos < len(ordered):
            return ordered[str_pos]
        return None
    except Exception:
        return None


def _get_region_label(imgt_pos, numbered: dict) -> str:
    """Return specific region label: CDR1, CDR2, CDR3, FR1, FR2, FR3, FR4."""
    try:
        cdr_by_region = numbered.get('cdr_by_region', {})
        for region, residues in cdr_by_region.items():
            if imgt_pos in residues:
                return region
        fr_by_region = numbered.get('fr_by_region', {})
        for region, residues in fr_by_region.items():
            if imgt_pos in residues:
                return region
        # Fallback: broad CDR/FR
        if imgt_pos in numbered['cdr_residues']:
            return 'CDR'
        return 'FR'
    except Exception:
        return 'unknown'


def find_liabilities(seq: str, numbered: Optional[dict] = None) -> dict:
    """
    Find sequence liability motifs with CDR/FR separation and IMGT position mapping.

    Args:
        seq:      Raw amino acid sequence string
        numbered: Output of number_sequence() for this sequence.
                  If provided, each flag is mapped to its IMGT position,
                  specific region (CDR1/2/3, FR1/2/3/4), and motif detail.
                  If None, falls back to string-position-only output
                  (backward compatible).

    Output per liability type:
        {liab}_cdr_count    int   — flags in CDR regions
        {liab}_fr_count     int   — flags in FR regions
        {liab}_detail       list  — one dict per flag:
            str_pos   int         string index (0-based)
            imgt_pos  int|tuple   IMGT position (None if mapping failed)
            region    str         CDR1/CDR2/CDR3/FR1/FR2/FR3/FR4
            motif     str         exact matched characters (e.g. 'NG', 'M')
    """

    def _scan(pattern: re.Pattern, seq: str, numbered: Optional[dict]) -> dict:
        flags = []
        for m in pattern.finditer(seq):
            str_pos = m.start()
            motif = m.group()
            imgt_pos = None
            region = 'unknown'

            if numbered is not None:
                mapping = _str_pos_to_imgt(str_pos, numbered)
                if mapping:
                    imgt_pos, broad = mapping
                    region = _get_region_label(imgt_pos, numbered)

            flags.append({
                'str_pos':  str_pos,
                'imgt_pos': imgt_pos,
                'region':   region,
                'motif':    motif,
            })

        cdr_flags = [f for f in flags if 'CDR' in f['region']]
        fr_flags = [f for f in flags if 'FR' in f['region']]

        return {
            'cdr_count': len(cdr_flags),
            'fr_count':  len(fr_flags),
            'detail':    flags,
        }

    # ── Run each liability scanner ────────────────────────────────────────────
    deam = _scan(DEAMIDATION_MOTIFS,   seq, numbered)
    oxid = _scan(OXIDATION_MOTIFS,     seq, numbered)
    isom = _scan(ISOMERIZATION_MOTIFS, seq, numbered)
    glyc = _scan(GLYCOSYLATION_MOTIF,  seq, numbered)
    asp = _scan(ASP_PRO_MOTIF,        seq, numbered)

    return {
        # Deamidation
        "deamidation_cdr_count":  deam['cdr_count'],
        "deamidation_fr_count":   deam['fr_count'],
        "deamidation_count":      deam['cdr_count'] + deam['fr_count'],
        "deamidation_detail":     str(deam['detail']),

        # Oxidation
        "oxidation_cdr_count":    oxid['cdr_count'],
        "oxidation_fr_count":     oxid['fr_count'],
        "oxidation_count":        oxid['cdr_count'] + oxid['fr_count'],
        "oxidation_detail":       str(oxid['detail']),

        # Isomerization
        "isomerization_cdr_count": isom['cdr_count'],
        "isomerization_fr_count":  isom['fr_count'],
        "isomerization_count":     isom['cdr_count'] + isom['fr_count'],
        "isomerization_detail":    str(isom['detail']),

        # N-glycosylation
        "glycosylation_cdr_count": glyc['cdr_count'],
        "glycosylation_fr_count":  glyc['fr_count'],
        "glycosylation_count":     glyc['cdr_count'] + glyc['fr_count'],
        "glycosylation_detail":    str(glyc['detail']),

        # Asp-Pro
        "asp_pro_cdr_count":      asp['cdr_count'],
        "asp_pro_fr_count":       asp['fr_count'],
        "asp_pro_count":          asp['cdr_count'] + asp['fr_count'],
        "asp_pro_detail":         str(asp['detail']),

        # Unpaired cysteines (whole sequence — not region-specific)
        "unpaired_cys":           max(0, seq.count('C') % 2),
    }


# ── Physicochemical properties ────────────────────────────────────────────────

def compute_physicochemical(seq: str) -> dict:
    """Compute pI, charge, MW, GRAVY, aromaticity for a sequence."""
    if not seq:
        return {}

    # Net charge at pH 7.4
    net_charge = sum(AA_CHARGE_PH7.get(aa, 0) for aa in seq)

    # Molecular weight
    mw = sum(AA_MW.get(aa, 111.1) for aa in seq) + 18.01  # +H2O

    # GRAVY (Grand Average of Hydropathicity)
    gravy = sum(AA_HYDROPHOBICITY.get(aa, 0) for aa in seq) / len(seq)

    # Aromaticity (fraction of F, Y, W)
    aromaticity = sum(1 for aa in seq if aa in 'FYW') / len(seq)

    # Simple pI estimation using iterative method
    pi = _estimate_pi(seq)

    # Amino acid composition
    aa_counts = {aa: seq.count(aa) for aa in 'ACDEFGHIKLMNPQRSTVWY'}

    return {
        "net_charge_ph7":  round(net_charge, 2),
        "mw_da":           round(mw, 1),
        "gravy":           round(gravy, 3),
        "aromaticity":     round(aromaticity, 3),
        "pi":              round(pi, 2),
        "length":          len(seq),
        **{f"aa_{aa}": count for aa, count in aa_counts.items()},
    }


def _estimate_pi(seq: str) -> float:
    """Estimate isoelectric point using iterative charge balance."""
    # pKa values for ionizable groups
    pka = {
        'D': 3.9, 'E': 4.1, 'H': 6.0, 'C': 8.3,
        'Y': 10.1, 'K': 10.5, 'R': 12.5,
        'Nterm': 8.0, 'Cterm': 3.1,
    }
    counts = {aa: seq.count(aa) for aa in 'DEHCYKR'}

    def charge_at_ph(ph):
        charge = (10**pka['Nterm']) / (10**pka['Nterm'] + 10**ph)
        charge -= (10**ph) / (10**pka['Cterm'] + 10**ph)
        for aa, pk, sign in [
            ('D', pka['D'], -1), ('E', pka['E'], -1),
            ('H', pka['H'], +1), ('C', pka['C'], -1),
            ('Y', pka['Y'], -1), ('K', pka['K'], +1), ('R', pka['R'], +1),
        ]:
            n = counts.get(aa, 0)
            if sign == +1:
                charge += n * (10**pk) / (10**pk + 10**ph)
            else:
                charge -= n * (10**ph) / (10**pk + 10**ph)
        return charge

    lo, hi = 0.0, 14.0
    for _ in range(100):
        mid = (lo + hi) / 2
        if charge_at_ph(mid) > 0:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2


# ── CDR properties ────────────────────────────────────────────────────────────

def compute_cdr_properties(numbered: dict) -> dict:
    """Compute CDR lengths, charge, hydrophobicity, and composition."""
    props = {}
    all_cdr_residues = []

    for region in ["CDR1", "CDR2", "CDR3"]:
        residues = numbered["cdr_by_region"].get(region, {})
        seq = "".join(residues[p] for p in sorted(
            residues.keys(),
            key=lambda x: (x[0], x[1]) if isinstance(x, tuple) else (x, ' ')
        ))
        length = len(seq)
        charge = sum(AA_CHARGE_PH7.get(aa, 0) for aa in seq)
        hydro = sum(AA_HYDROPHOBICITY.get(aa, 0)
                    for aa in seq) / length if length else 0

        props[f"{region.lower()}_length"] = length
        props[f"{region.lower()}_charge"] = round(charge, 2)
        props[f"{region.lower()}_hydrophobicity"] = round(hydro, 3)
        props[f"{region.lower()}_sequence"] = seq
        all_cdr_residues.extend(seq)

    props["total_cdr_length"] = sum(
        props[f"{r}_length"] for r in ["cdr1", "cdr2", "cdr3"]
    )
    if all_cdr_residues:
        props["cdr_pct_hydrophobic"] = round(
            sum(1 for aa in all_cdr_residues if AA_HYDROPHOBICITY.get(aa, 0) > 0)
            / len(all_cdr_residues), 3
        )
        props["cdr_pct_charged"] = round(
            sum(1 for aa in all_cdr_residues if aa in 'RKHDE')
            / len(all_cdr_residues), 3
        )

    return props


# ── Vernier zone analysis ─────────────────────────────────────────────────────

def compute_vernier(
    query_numbered:   dict,
    mouse_numbered:   dict,
    grafted_numbered: Optional[dict],
    chain_type:       str,
) -> dict:
    """
    Analyse back-mutations at FR positions, with Vernier zone subgrouping.

    Correct three-way logic per FR position:

        mouse_aa == grafted_aa  → no mutation needed (germlines already agree)
                                  not counted as a back-mutation candidate

        mouse_aa != grafted_aa  → this position was humanized during grafting
                                  = a "mutable" position

            query_aa == mouse_aa   → mouse residue was RESTORED → back-mutated
            query_aa == grafted_aa → human residue was KEPT     → humanized
            query_aa != either     → substituted to a third residue

    Each mutable position is also labelled as Vernier or non-Vernier.

    Args:
        query_numbered:   the humanized sequence being evaluated (any of seqs 1–9)
        mouse_numbered:   the original mouse sequence
        grafted_numbered: the CDR-grafted baseline BEFORE back-mutations
                          (seq 3 for lab pipeline, seq 1 for pipeline grafted,
                           seq 4 for detected grafted, seq 8 for stated grafted)
                          If None, grafted baseline is assumed identical to mouse
                          (all FR positions treated as mutable).
        chain_type:       "H" for VH, "K"/"L" for VL

    Output columns (all prefixed by caller with vh_ or vl_):

      Back-mutation summary (all FR positions):
        fr_mutable_count          total FR positions where mouse != grafted
        fr_backmut_count          positions restored to mouse
        fr_humanized_count        positions kept as human germline
        fr_other_count            positions changed to a third residue

      Vernier subgroup:
        vernier_mutable_count     mutable positions that are Vernier
        vernier_backmut_count     Vernier positions restored to mouse
        vernier_humanized_count   Vernier positions kept as human
        vernier_other_count       Vernier positions changed to third residue

      Non-Vernier subgroup:
        non_vernier_mutable_count
        non_vernier_backmut_count
        non_vernier_humanized_count
        non_vernier_other_count

      Per-position detail (as string-serialised list of dicts):
        backmut_detail    all mutable positions with full residue info
    """
    vernier_set = VERNIER_VH if chain_type == "H" else VERNIER_VL

    query_fr = query_numbered["fr_residues"]
    mouse_fr = mouse_numbered["fr_residues"]

    # seq 0/0r: no grafting baseline — Sapiens ran directly on mouse.
    # Every FR position where query != mouse was changed by Sapiens.
    # There is no "mutable" concept (no humanization step preceded Sapiens),
    # so all positions where query != mouse are labelled "humanized_by_sapiens".
    seq0_mode = grafted_numbered is None

    grafted_fr = grafted_numbered["fr_residues"] if grafted_numbered else {}

    # All FR positions present in both mouse and query
    all_fr_positions = set(mouse_fr) & set(query_fr)

    detail = []
    for pos in sorted(all_fr_positions):
        mouse_aa = mouse_fr.get(pos)
        query_aa = query_fr.get(pos)

        if mouse_aa is None or query_aa is None:
            continue

        if seq0_mode:
            # No grafting baseline — every position where Sapiens
            # changed from mouse is recorded as "humanized_by_sapiens"
            if mouse_aa == query_aa:
                continue  # Sapiens kept mouse residue — not interesting
            status = "humanized_by_sapiens"
            grafted_aa = None  # no grafted baseline exists
        else:
            grafted_aa = grafted_fr.get(pos)
            if grafted_aa is None:
                continue
            # Skip positions where mouse and grafted already agree — not mutable
            if mouse_aa == grafted_aa:
                continue
            # Three-way classification
            if query_aa == mouse_aa:
                status = "back_mutated"
            elif query_aa == grafted_aa:
                status = "humanized"
            else:
                status = "other_substitution"

        is_vernier = pos in vernier_set

        detail.append({
            "imgt_pos":   pos,
            "is_vernier": is_vernier,
            "mouse_aa":   mouse_aa,
            "grafted_aa": grafted_aa,
            "query_aa":   query_aa,
            "status":     status,
        })

    # ── Aggregate counts ──────────────────────────────────────────────────────
    def counts(subset):
        return {
            "mutable":   len(subset),
            "backmut":   sum(1 for d in subset if d["status"] == "back_mutated"),
            "humanized": sum(1 for d in subset if d["status"] == "humanized"),
            "other":     sum(1 for d in subset if d["status"] == "other_substitution"),
        }

    all_c = counts(detail)
    vern_c = counts([d for d in detail if d["is_vernier"]])
    nonvern_c = counts([d for d in detail if not d["is_vernier"]])

    return {
        # All FR mutable positions
        "fr_mutable_count":           all_c["mutable"],
        "fr_backmut_count":           all_c["backmut"],
        "fr_humanized_count":         all_c["humanized"],
        "fr_other_count":             all_c["other"],

        # Vernier subgroup
        "vernier_mutable_count":      vern_c["mutable"],
        "vernier_backmut_count":      vern_c["backmut"],
        "vernier_humanized_count":    vern_c["humanized"],
        "vernier_other_count":        vern_c["other"],

        # Non-Vernier subgroup
        "non_vernier_mutable_count":  nonvern_c["mutable"],
        "non_vernier_backmut_count":  nonvern_c["backmut"],
        "non_vernier_humanized_count": nonvern_c["humanized"],
        "non_vernier_other_count":    nonvern_c["other"],

        # Full per-position detail
        "backmut_detail":             str(detail),
    }


# ── OASis humanness ───────────────────────────────────────────────────────────

OASIS_DB = "/workspace/antibody-humanization-tool/data/OASis_9mers_v1.db"
OASIS_THRESHOLD = 0.10   # matches BioPhi CLI default: --min-percent-subjects 10


def _safe_round(val, ndigits: int = 4) -> Optional[float]:
    """Safely round a value, returning None if not numeric."""
    try:
        return round(float(val), ndigits) if val is not None else None
    except (TypeError, ValueError):
        return None


def compute_oasis_per_position(
    seq:        str,
    numbered:   Optional[dict] = None,
    chain_type: str = "H",
) -> dict:
    """
    Compute per-position OASis humanness using BioPhi's own functions directly.

    Uses get_chain_humanness() from biophi.humanization.methods.humanness —
    the same function called internally by the BioPhi CLI — guaranteeing
    identical results to `biophi oasis` output.

    BioPhi's logic:
      - Denominator = subjects with >= 10,000 complete Heavy/Light seqs,
        excluding Corcoran_2016 study (chain-specific)
      - Human threshold = 10% of qualifying subjects (CLI default)
      - 9-mer window slides across the full numbered sequence

    Args:
        seq:        raw amino acid sequence string
        numbered:   output of number_sequence() — used for CDR/FR region labels
                    on per-position detail (optional)
        chain_type: "H" for VH, "K" or "L" for VL

    Returns (prefixed by caller with vh_ or vl_):
        oasis_per_position_detail   str(list) one dict per 9-mer:
            str_pos, imgt_pos, region, nmer, n_subjects, prevalence, is_human
        oasis_identity              float  matches BioPhi CLI output exactly
        oasis_n_human_nmers         int
        oasis_n_nonhuman_nmers      int
        oasis_n_total_subjects      int    chain-specific denominator
        oasis_cdr_identity          float  identity for CDR-start 9-mers
        oasis_fr_identity           float  identity for FR-start 9-mers
        oasis_nonhuman_detail       str(list) non-human 9-mers for viz
    """
    empty = {
        "oasis_per_position_detail": str([]),
        "oasis_identity":            None,
        "oasis_n_human_nmers":       None,
        "oasis_n_nonhuman_nmers":    None,
        "oasis_n_total_subjects":    None,
        "oasis_cdr_identity":        None,
        "oasis_fr_identity":         None,
        "oasis_nonhuman_detail":     str([]),
    }

    if len(seq) < 9:
        return empty

    try:
        from biophi.humanization.methods.humanness import (
            get_chain_humanness, OASisParams
        )
        from abnumber import Chain

        # BioPhi expects an abnumber Chain object
        # chain_type is inferred automatically from the sequence
        chain = Chain(seq, scheme="imgt", cdr_definition="imgt",
                      assign_germline=True)

        params = OASisParams(
            oasis_db_path=OASIS_DB,
            min_fraction_subjects=OASIS_THRESHOLD,
        )

        # Call BioPhi's own function — identical to what the CLI does
        chain_humanness = get_chain_humanness(chain, params=params)

        # ── Extract per-position detail ───────────────────────────────────────
        # chain_humanness.peptides is {Position: PeptideHumanness}
        # Position objects have an IMGT position number
        results = []
        for i, (pos, peptide) in enumerate(chain_humanness.peptides.items()):
            # BioPhi Position objects render as e.g. "111" or "111A" for insertions
            # Extract numeric part only — insertion letter is not needed here
            try:
                pos_str = str(pos).strip()
                # Remove chain prefix if present (H111A → 111A)
                pos_str = pos_str.lstrip("HLK")
                # Extract leading integer — covers "111" and "111A"
                num_part = "".join(c for c in pos_str if c.isdigit())
                imgt_pos = int(num_part) if num_part else None
            except Exception:
                imgt_pos = None

            # Get CDR/FR region label from our numbering if available
            region = "unknown"
            if numbered is not None and imgt_pos is not None:
                region = _get_region_label(imgt_pos, numbered)

            is_human = bool(peptide.is_human(OASIS_THRESHOLD))

            results.append({
                "str_pos":    i,
                "imgt_pos":   imgt_pos,
                "region":     region,
                "nmer":       peptide.seq,
                "n_subjects": peptide.num_oas_subjects,
                "prevalence": round(peptide.fraction_oas_subjects, 4)
                if peptide.fraction_oas_subjects is not None else None,
                "is_human":   is_human,
            })

        # ── Aggregates ────────────────────────────────────────────────────────
        n_total = chain_humanness.get_num_peptides()
        n_human = chain_humanness.get_num_human_peptides(OASIS_THRESHOLD)
        n_nonhuman = n_total - n_human
        identity = round(
            chain_humanness.get_oasis_identity(OASIS_THRESHOLD), 4)

        cdr_results = [r for r in results if "CDR" in r["region"]]
        fr_results = [r for r in results if "FR" in r["region"]]

        cdr_identity = (
            round(
                sum(1 for r in cdr_results if r["is_human"]) / len(cdr_results), 4)
            if cdr_results else None
        )
        fr_identity = (
            round(
                sum(1 for r in fr_results if r["is_human"]) / len(fr_results), 4)
            if fr_results else None
        )

        # n_total_subjects: derive from denominator used by BioPhi
        # (fraction_oas_subjects = n_subjects / n_total_subjects)
        # Back-calculate from first peptide with known subjects
        n_total_subjects = None
        for peptide in chain_humanness.peptides.values():
            if peptide.num_oas_subjects and peptide.fraction_oas_subjects:
                n_total_subjects = round(
                    peptide.num_oas_subjects / peptide.fraction_oas_subjects
                )
                break

        return {
            "oasis_per_position_detail": str(results),
            "oasis_identity":            identity,
            "oasis_n_human_nmers":       n_human,
            "oasis_n_nonhuman_nmers":    n_nonhuman,
            "oasis_n_total_subjects":    n_total_subjects,
            "oasis_cdr_identity":        cdr_identity,
            "oasis_fr_identity":         fr_identity,
            "oasis_nonhuman_detail":     str([r for r in results if not r["is_human"]]),
        }

    except Exception as e:
        print(f"    OASis per-position failed: {e}")
        return empty


def compute_oasis(vh_seq: str, vl_seq: str, clone_id: str = "",
                  seq_id: str = "") -> dict:
    """
    Compute OASis humanness score via BioPhi CLI.
    Writes a temporary FASTA, runs biophi oasis, parses the XLSX output.

    Returns:
        oasis_vh:            OASis humanness score for VH (0-1)
        oasis_vl:            OASis humanness score for VL (0-1)
        oasis_pct_human_vh:  % 9-mers found in human repertoire (VH)
        oasis_pct_human_vl:  % 9-mers found in human repertoire (VL)
    """
    empty = {
        "oasis_identity": None, "oasis_percentile": None,
        "oasis_germline_content": None,
        "oasis_vh_identity": None, "oasis_vh_percentile": None,
        "oasis_vh_germline": None, "oasis_vh_nonhuman_count": None,
        "oasis_vh_v_germline": None, "oasis_vh_j_germline": None,
        "oasis_vl_identity": None, "oasis_vl_percentile": None,
        "oasis_vl_germline": None, "oasis_vl_nonhuman_count": None,
        "oasis_vl_v_germline": None, "oasis_vl_j_germline": None,
    }
    try:
        import subprocess
        import tempfile
        import os
        import openpyxl

        tag = f"{clone_id}_{seq_id}".replace("/", "_")

        with tempfile.TemporaryDirectory() as tmpdir:
            fasta_path = os.path.join(tmpdir, f"{tag}.fa")
            output_path = os.path.join(tmpdir, f"{tag}.xlsx")

            # Write paired FASTA — BioPhi expects VH and VL as separate entries
            with open(fasta_path, "w") as f:
                if vh_seq:
                    f.write(f">{tag}_VH\n{vh_seq}\n")
                if vl_seq:
                    f.write(f">{tag}_VL\n{vl_seq}\n")

            # Run BioPhi OASis CLI
            result = subprocess.run(
                [
                    "biophi", "oasis", fasta_path,
                    "--output",   output_path,
                    "--oasis-db", OASIS_DB,
                    "--scheme",   "imgt",
                    "--summary",
                ],
                capture_output=True, text=True, timeout=120
            )

            if result.returncode != 0:
                print(f"    OASis CLI error: {result.stderr.strip()[:200]}")
                return empty

            if not os.path.exists(output_path):
                print(f"    OASis: no output file generated")
                return empty

            # Parse XLSX summary sheet
            wb = openpyxl.load_workbook(output_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows())]

            scores = {"oasis_vh": None, "oasis_vl": None,
                      "oasis_pct_human_vh": None, "oasis_pct_human_vl": None}

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                # Column names from BioPhi 1.0.11 CLI --summary output:
                # Antibody, OASis Identity, Heavy OASis Identity,
                # Light OASis Identity, Heavy Germline Content,
                # Light Germline Content, Heavy Non-human peptides, etc.
                scores = {
                    "oasis_identity":          _safe_round(row_dict.get("OASis Identity")),
                    "oasis_percentile":        _safe_round(row_dict.get("OASis Percentile")),
                    "oasis_germline_content":  _safe_round(row_dict.get("Germline Content")),
                    "oasis_vh_identity":       _safe_round(row_dict.get("Heavy OASis Identity")),
                    "oasis_vh_percentile":     _safe_round(row_dict.get("Heavy OASis Percentile")),
                    "oasis_vh_germline":       _safe_round(row_dict.get("Heavy Germline Content")),
                    "oasis_vh_nonhuman_count": row_dict.get("Heavy Non-human peptides"),
                    "oasis_vh_v_germline":     row_dict.get("Heavy V Germline"),
                    "oasis_vh_j_germline":     row_dict.get("Heavy J Germline"),
                    "oasis_vl_identity":       _safe_round(row_dict.get("Light OASis Identity")),
                    "oasis_vl_percentile":     _safe_round(row_dict.get("Light OASis Percentile")),
                    "oasis_vl_germline":       _safe_round(row_dict.get("Light Germline Content")),
                    "oasis_vl_nonhuman_count": row_dict.get("Light Non-human peptides"),
                    "oasis_vl_v_germline":     row_dict.get("Light V Germline"),
                    "oasis_vl_j_germline":     row_dict.get("Light J Germline"),
                }
                break  # only one row in summary mode

            return scores

    except FileNotFoundError:
        print("    OASis: biophi CLI not found")
        return empty
    except Exception as e:
        print(f"    OASis failed: {e}")
        return empty


# ── Germline identity ─────────────────────────────────────────────────────────

def compute_germline_identity(
    query_numbered: dict,
    germline_name: str,
    chain_type: str,
) -> Optional[float]:
    """Compute FR identity between query sequence and named germline."""
    if not germline_name:
        return None
    try:
        from pipeline.step_b_germline_scoring import rank_germlines
        rankings = rank_germlines(
            query_fr=query_numbered["fr_residues"],
            chain_type=chain_type,
            top_n=1,
            min_identity=0.0,
        )
        # Find this specific germline in rankings
        from pipeline.step_b_germline_scoring import _GERMLINE_FR_DB
        from pipeline.step_a_numbering import ALL_FR_POSITIONS
        germ_gene = germline_name.split("*")[0]
        germ_fr = {}
        for allele, residues in _GERMLINE_FR_DB.get(chain_type, {}).items():
            if allele.split("*")[0] == germ_gene:
                germ_fr = residues
                break
        if not germ_fr:
            return None
        query_fr = query_numbered["fr_residues"]
        matched = comparable = 0
        for pos in ALL_FR_POSITIONS:
            q = query_fr.get(pos)
            g = germ_fr.get(pos)
            if q and g:
                comparable += 1
                if q == g:
                    matched += 1
        return round(matched / comparable, 3) if comparable else None
    except Exception:
        return None


# ── Structure scoring ─────────────────────────────────────────────────────────

# ── CamSol intrinsic solubility ───────────────────────────────────────────────
# Reimplementation of the CamSol intrinsic (sequence-only) algorithm.
# Reference: Sormanni et al., J. Mol. Biol. 2015
#            doi:10.1016/j.jmb.2014.09.026
#
# Algorithm:
#   1. Per-residue hydrophobicity (Wimley-White scale, sign-inverted so
#      positive = aggregation-prone, negative = solubility-promoting)
#   2. Per-residue charge at pH 7
#   3. Secondary structure propensity correction (beta-sheet propensity
#      increases aggregation risk; alpha-helix propensity reduces it)
#   4. Smoothing over a window of 5 neighbors (window size 9)
#   5. Final score = mean of smoothed profile = intrinsic solubility score
#      Higher score = more soluble

# Wimley-White hydrophobicity scale (sign-inverted from CamSol convention:
# positive = hydrophobic = aggregation-prone)
_CAMSOL_HYDROPHOBICITY = {
    'F':  1.06, 'I':  0.67, 'L':  0.57, 'W':  0.50, 'V':  0.40,
    'M':  0.26, 'A':  0.17, 'C':  0.13, 'P': -0.29, 'G': -0.01,
    'T': -0.40, 'S': -0.50, 'Y': -0.41, 'H': -0.96, 'Q': -0.58,
    'N': -0.82, 'K': -1.35, 'D': -1.23, 'E': -1.31, 'R': -1.37,
}

# Charge at pH 7.4
_CAMSOL_CHARGE = {
    'R': +1.0, 'K': +1.0, 'H': +0.1,
    'D': -1.0, 'E': -1.0,
}

# Beta-sheet propensity (positive = promotes aggregation)
# Alpha-helix propensity (negative = opposes aggregation)
# Combined as: beta - alpha, from Sormanni 2015 Table S2 approximation
_CAMSOL_STRUCT_PROPENSITY = {
    'V':  0.60, 'I':  0.50, 'F':  0.40, 'T':  0.35, 'W':  0.30,
    'L':  0.25, 'Y':  0.20, 'C':  0.15, 'M':  0.05, 'A': -0.20,
    'G': -0.10, 'S': -0.20, 'H': -0.25, 'Q': -0.30, 'N': -0.35,
    'P': -0.50, 'D': -0.30, 'E': -0.40, 'K': -0.45, 'R': -0.40,
}


def compute_structure_scores(vh_seq: str, vl_seq: str, clone_id: str) -> dict:
    """
    Predict Fv structure with ABodyBuilder2 and compute per-residue confidence scores.

    ABodyBuilder2 uses an ensemble of 4 models and outputs error_estimates —
    lower error = higher confidence (inverse of pLDDT).
    We use the best-ranked model and convert to a 0-100 confidence scale:
        confidence = 100 * exp(-error_estimate)
    This maps:
        error=0.0  → confidence=100  (perfect)
        error=0.1  → confidence=90
        error=0.5  → confidence=61
        error=1.0  → confidence=37
        error=2.0  → confidence=14

    Requires: pip install ImmuneBuilder --break-system-packages
    Runs without OpenMM/pdbfixer (no structure refinement).
    """
    empty = {
        "conf_mean_vh": None, "conf_mean_vl": None, "conf_mean_fv": None,
        "conf_min_vh":  None, "conf_min_vl":  None,
        "conf_cdr_mean_vh": None, "conf_cdr_mean_vl": None,
        "conf_fr_mean_vh":  None, "conf_fr_mean_vl":  None,
    }

    try:
        import numpy as np
        from ImmuneBuilder import ABodyBuilder2

        predictor = ABodyBuilder2()
        antibody = predictor.predict({"H": vh_seq, "L": vl_seq})

        # error_estimates shape: [4 models, n_residues]
        # ranking: best model first
        best_model_idx = antibody.ranking[0]
        errors = antibody.error_estimates[best_model_idx].cpu().numpy()

        # Convert error → confidence score 0-100
        confidence = 100.0 * np.exp(-errors)

        # Split VH and VL by sequence length
        n_vh = len(vh_seq)
        conf_vh = confidence[:n_vh]
        conf_vl = confidence[n_vh:n_vh + len(vl_seq)]

        # Map to CDR/FR regions using our numbering
        try:
            vh_num = number_sequence(vh_seq, chain_type="H")
            vl_num = number_sequence(vl_seq, chain_type=None)

            def _imgt_sort_key(pos):
                """Sort IMGT positions including insertion tuples e.g. (111,'A')."""
                if isinstance(pos, tuple):
                    return (pos[0], pos[1])
                return (pos, ' ')

            def split_cdr_fr(num, conf_arr):
                """Split confidence scores into CDR and FR arrays."""
                all_res = {}
                for pos in num["fr_residues"]:
                    all_res[pos] = "FR"
                for pos in num["cdr_residues"]:
                    all_res[pos] = "CDR"
                ordered = sorted(all_res.keys(), key=_imgt_sort_key)
                cdr_conf = [conf_arr[i] for i, pos in enumerate(ordered)
                            if i < len(conf_arr) and all_res[pos] == "CDR"]
                fr_conf = [conf_arr[i] for i, pos in enumerate(ordered)
                           if i < len(conf_arr) and all_res[pos] == "FR"]
                return np.array(cdr_conf), np.array(fr_conf)

            vh_cdr_conf, vh_fr_conf = split_cdr_fr(vh_num, conf_vh)
            vl_cdr_conf, vl_fr_conf = split_cdr_fr(vl_num, conf_vl)

            cdr_mean_vh = round(float(np.mean(vh_cdr_conf)),
                                2) if len(vh_cdr_conf) else None
            cdr_mean_vl = round(float(np.mean(vl_cdr_conf)),
                                2) if len(vl_cdr_conf) else None
            fr_mean_vh = round(float(np.mean(vh_fr_conf)),
                               2) if len(vh_fr_conf) else None
            fr_mean_vl = round(float(np.mean(vl_fr_conf)),
                               2) if len(vl_fr_conf) else None
        except Exception:
            cdr_mean_vh = cdr_mean_vl = fr_mean_vh = fr_mean_vl = None

        return {
            "conf_mean_vh":     round(float(np.mean(conf_vh)), 2),
            "conf_mean_vl":     round(float(np.mean(conf_vl)), 2),
            "conf_mean_fv":     round(float(np.mean(confidence)), 2),
            "conf_min_vh":      round(float(np.min(conf_vh)),  2),
            "conf_min_vl":      round(float(np.min(conf_vl)),  2),
            "conf_cdr_mean_vh": cdr_mean_vh,
            "conf_cdr_mean_vl": cdr_mean_vl,
            "conf_fr_mean_vh":  fr_mean_vh,
            "conf_fr_mean_vl":  fr_mean_vl,
        }

    except ImportError:
        print(f"    ABodyBuilder2 not available — pip install ImmuneBuilder")
        return empty
    except Exception as e:
        print(f"    Structure scoring failed for {clone_id}: {e}")
        return empty


# ── Sequence identity between two sequences ───────────────────────────────────

def sequence_identity(seq_a: str, seq_b: str, chain_type: str) -> Optional[float]:
    """Compute FR+CDR sequence identity between two numbered sequences."""
    if not seq_a or not seq_b:
        return None
    try:
        num_a = number_sequence(seq_a, chain_type=chain_type)
        num_b = number_sequence(seq_b, chain_type=chain_type)
        all_a = {**num_a["fr_residues"], **{k: v for k,
                                            v in num_a["cdr_residues"].items() if isinstance(k, int)}}
        all_b = {**num_b["fr_residues"], **{k: v for k,
                                            v in num_b["cdr_residues"].items() if isinstance(k, int)}}
        positions = set(all_a) & set(all_b)
        if not positions:
            return None
        matched = sum(1 for p in positions if all_a[p] == all_b[p])
        return round(matched / len(positions), 3)
    except Exception:
        return None


# ── Main scoring function ─────────────────────────────────────────────────────

def compute_camsol(seq: str, numbered: Optional[dict] = None) -> dict:
    """
    Compute CamSol intrinsic solubility score from sequence alone.

    Per-residue score combines hydrophobicity, charge, and secondary
    structure propensity, then applies a smoothing window.

    Higher score = more soluble (less aggregation-prone).
    Typical range for antibody variable domains: -1.5 to +1.5.

    Args:
        seq:      raw amino acid sequence string
        numbered: output of number_sequence() for CDR/FR region labeling

    Returns:
        camsol_score          float  overall mean solubility score
        camsol_vh_cdr_score   float  mean score restricted to CDR residues
        camsol_vh_fr_score    float  mean score restricted to FR residues
        camsol_per_residue    str    per-residue scores as list of dicts:
            str_pos, imgt_pos, region, aa, raw_score, smoothed_score
        camsol_hotspot_count  int    residues with smoothed_score < -0.5
                                     (aggregation hotspots)
    """
    empty = {
        "camsol_score":         None,
        "camsol_cdr_score":     None,
        "camsol_fr_score":      None,
        "camsol_per_residue":   str([]),
        "camsol_hotspot_count": None,
    }

    if not seq or len(seq) < 3:
        return empty

    try:
        import numpy as np

        n = len(seq)

        # ── Step 1: per-residue raw score ─────────────────────────────────────
        raw = np.zeros(n)
        for i, aa in enumerate(seq):
            hydro = _CAMSOL_HYDROPHOBICITY.get(aa, 0.0)
            charge = _CAMSOL_CHARGE.get(aa, 0.0)
            struct = _CAMSOL_STRUCT_PROPENSITY.get(aa, 0.0)
            # CamSol raw score: hydrophobicity + structural propensity
            # charge acts as a solubility promoter (negative of abs charge
            # for charged residues — charged residues are generally soluble)
            raw[i] = hydro + struct - abs(charge) * 0.5

        # ── Step 2: smoothing window of ±4 neighbors (window size 9) ─────────
        window = 9
        half = window // 2
        smoothed = np.zeros(n)
        for i in range(n):
            start = max(0, i - half)
            end = min(n, i + half + 1)
            smoothed[i] = np.mean(raw[start:end])

        # ── Step 3: map to IMGT positions and regions ─────────────────────────
        results = []
        for i, (aa, s_score, r_score) in enumerate(zip(seq, smoothed, raw)):
            imgt_pos = None
            region = "unknown"
            if numbered is not None:
                mapping = _str_pos_to_imgt(i, numbered)
                if mapping:
                    imgt_pos, _ = mapping
                    region = _get_region_label(imgt_pos, numbered)

            results.append({
                "str_pos":       i,
                "imgt_pos":      imgt_pos,
                "region":        region,
                "aa":            aa,
                "raw_score":     round(float(r_score), 4),
                "smoothed_score": round(float(s_score), 4),
            })

        # ── Step 4: aggregates ────────────────────────────────────────────────
        overall_score = round(float(np.mean(smoothed)), 4)

        cdr_scores = [r["smoothed_score"]
                      for r in results if "CDR" in r["region"]]
        fr_scores = [r["smoothed_score"]
                     for r in results if "FR" in r["region"]]

        cdr_score = round(float(np.mean(cdr_scores)),
                          4) if cdr_scores else None
        fr_score = round(float(np.mean(fr_scores)),  4) if fr_scores else None

        # Aggregation hotspots: smoothed score below -0.5
        hotspot_count = sum(1 for r in results if r["smoothed_score"] < -0.5)

        return {
            "camsol_score":         overall_score,
            "camsol_cdr_score":     cdr_score,
            "camsol_fr_score":      fr_score,
            "camsol_per_residue":   str(results),
            "camsol_hotspot_count": hotspot_count,
        }

    except Exception as e:
        print(f"    CamSol failed: {e}")
        return empty


def score_one(
    clone_id:     str,
    seq_id:       str,
    seq_label:    str,
    vh_seq:       str,
    vl_seq:       str,
    vl_chain_type: str,
    mouse_vh:     str,
    mouse_vl:     str,
    lab_hu_vh:    str,
    lab_hu_vl:    str,
    lab_final_vh: str,
    lab_final_vl: str,
    vh_germline:  str,
    vl_germline:  str,
    grafted_vh:   str = "",
    grafted_vl:   str = "",
    run_structure: bool = False,
) -> dict:
    """Score a single sequence entry.

    grafted_vh/vl: the CDR-grafted baseline sequence for this seq_id.
      Used by compute_vernier() to correctly classify back-mutations.
      Correct pairing:
        seq 0, 0r  → mouse sequence  (Sapiens ran on mouse directly)
        seq 1      → seq 1 itself    (IS the grafted baseline)
        seq 2, 2r  → seq 1
        seq 3      → seq 3 itself
        seq 4      → seq 4 itself
        seq 5      → seq 3
        seq 6, 6r  → seq 4
        seq 7      → seq 4
        seq 8      → seq 8 itself
        seq 9, 9r  → seq 8
    """
    row = {
        "clone":     clone_id,
        "seq_id":    seq_id,
        "seq_label": seq_label,
    }

    # Number all sequences needed
    try:
        vh_num = number_sequence(
            vh_seq,        chain_type="H") if vh_seq else None
        vl_num = number_sequence(
            vl_seq,        chain_type=vl_chain_type) if vl_seq else None
        mouse_vh_num = number_sequence(
            mouse_vh,      chain_type="H") if mouse_vh else None
        mouse_vl_num = number_sequence(
            mouse_vl,      chain_type=vl_chain_type) if mouse_vl else None
        lab_hu_vh_num = number_sequence(
            lab_hu_vh,     chain_type="H") if lab_hu_vh else None
        lab_hu_vl_num = number_sequence(
            lab_hu_vl,     chain_type=vl_chain_type) if lab_hu_vl else None
        lab_fin_vh_num = number_sequence(
            lab_final_vh,  chain_type="H") if lab_final_vh else None
        lab_fin_vl_num = number_sequence(
            lab_final_vl,  chain_type=vl_chain_type) if lab_final_vl else None
        # Correct grafted baseline for this seq_id (passed explicitly from main)
        grafted_vh_num = number_sequence(
            grafted_vh,    chain_type="H") if grafted_vh else None
        grafted_vl_num = number_sequence(
            grafted_vl,    chain_type=vl_chain_type) if grafted_vl else None
    except Exception as e:
        print(f"    Numbering failed for {clone_id} seq{seq_id}: {e}")
        return row

    # ── VH features ───────────────────────────────────────────────────────────
    if vh_seq and vh_num:
        # Physicochemical
        row.update(
            {f"vh_{k}": v for k, v in compute_physicochemical(vh_seq).items()})

        # CDR properties
        row.update(
            {f"vh_{k}": v for k, v in compute_cdr_properties(vh_num).items()})

        # Sequence liabilities (CDR/FR separated, IMGT positions mapped)
        row.update(
            {f"vh_{k}": v for k, v in find_liabilities(vh_seq, vh_num).items()})

        # Germline identity
        # For seq 0/0r: detect closest germline from actual FR residues
        # since vh_germline in CSV is the pipeline germline (used for seq 1),
        # not a germline that Sapiens selected or converged toward.
        if seq_id in ("0", "0r"):
            try:
                from pipeline.step_b_germline_scoring import rank_germlines
                vh_rankings = rank_germlines(
                    vh_num["fr_residues"], "H", top_n=1)
                detected_vh_germ = vh_rankings[0]["germline"] if vh_rankings else None
            except Exception:
                detected_vh_germ = None
            row["vh_detected_germline_seq0"] = detected_vh_germ
            row["vh_germline_identity"] = compute_germline_identity(
                vh_num, detected_vh_germ, "H")
        else:
            row["vh_germline_identity"] = compute_germline_identity(
                vh_num, vh_germline, "H")

        # Sequence identity vs lab sequences
        row["vh_identity_vs_lab_hu"] = sequence_identity(
            vh_seq, lab_hu_vh,    "H")
        row["vh_identity_vs_lab_final"] = sequence_identity(
            vh_seq, lab_final_vh, "H")
        row["vh_identity_vs_mouse"] = sequence_identity(
            vh_seq, mouse_vh,     "H")

        # OASis per-position humanness (direct SQLite query)
        row.update({f"vh_{k}": v for k, v in
                    compute_oasis_per_position(vh_seq, vh_num, "H").items()})

        # CamSol intrinsic solubility
        row.update({f"vh_{k}": v for k, v in
                    compute_camsol(vh_seq, vh_num).items()})

        # Vernier zone — back-mutation analysis using correct grafted baseline
        if mouse_vh_num:
            row.update({f"vh_{k}": v for k, v in compute_vernier(
                vh_num, mouse_vh_num, grafted_vh_num, "H"
            ).items()})

    # ── VL features ───────────────────────────────────────────────────────────
    if vl_seq and vl_num:
        row.update(
            {f"vl_{k}": v for k, v in compute_physicochemical(vl_seq).items()})
        row.update(
            {f"vl_{k}": v for k, v in compute_cdr_properties(vl_num).items()})
        row.update(
            {f"vl_{k}": v for k, v in find_liabilities(vl_seq, vl_num).items()})

        # Germline identity — detect for seq 0/0r, use stored for others
        if seq_id in ("0", "0r"):
            try:
                from pipeline.step_b_germline_scoring import rank_germlines
                vl_rankings = rank_germlines(
                    vl_num["fr_residues"], vl_chain_type, top_n=1)
                detected_vl_germ = vl_rankings[0]["germline"] if vl_rankings else None
            except Exception:
                detected_vl_germ = None
            row["vl_detected_germline_seq0"] = detected_vl_germ
            row["vl_germline_identity"] = compute_germline_identity(
                vl_num, detected_vl_germ, vl_chain_type)
        else:
            row["vl_germline_identity"] = compute_germline_identity(
                vl_num, vl_germline, vl_chain_type)

        row["vl_identity_vs_lab_hu"] = sequence_identity(
            vl_seq, lab_hu_vl,    vl_chain_type)
        row["vl_identity_vs_lab_final"] = sequence_identity(
            vl_seq, lab_final_vl, vl_chain_type)
        row["vl_identity_vs_mouse"] = sequence_identity(
            vl_seq, mouse_vl,     vl_chain_type)

        # OASis per-position humanness (direct SQLite query)
        row.update({f"vl_{k}": v for k, v in
                    compute_oasis_per_position(vl_seq, vl_num, vl_chain_type).items()})

        # CamSol intrinsic solubility
        row.update({f"vl_{k}": v for k, v in
                    compute_camsol(vl_seq, vl_num).items()})

        if mouse_vl_num:
            row.update({f"vl_{k}": v for k, v in compute_vernier(
                vl_num, mouse_vl_num, grafted_vl_num, vl_chain_type
            ).items()})

    # ── Paired Fv features ────────────────────────────────────────────────────
    if vh_seq and vl_seq:
        fv_seq = vh_seq + vl_seq
        row["fv_length"] = len(fv_seq)
        row["fv_net_charge_ph7"] = round(
            sum(AA_CHARGE_PH7.get(aa, 0) for aa in fv_seq), 2)
        row["fv_pi"] = round(_estimate_pi(fv_seq), 2)

        # OASis (paired)
        oasis = compute_oasis(vh_seq, vl_seq, clone_id=clone_id, seq_id=seq_id)
        row.update(oasis)

        # Structure (optional)
        if run_structure:
            struct = compute_structure_scores(vh_seq, vl_seq, clone_id)
            row.update(struct)

    return row


# ── CSV loaders ───────────────────────────────────────────────────────────────

def load_generated(csv_path: str) -> list[dict]:
    """Load generated sequences CSV."""
    rows = []
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, dialect="excel"):
            rows.append({k: v.strip() for k, v in row.items()})
    return rows


def load_benchmark(csv_path: str) -> dict:
    """Load benchmark CSV. Returns {clone_id: row_dict}."""
    benchmarks = {}
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, dialect="excel"):
            clone = row.get("clone", "").strip()
            if clone:
                benchmarks[clone] = {k: v.strip() for k, v in row.items()}
    return benchmarks


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Score generated sequences")
    parser.add_argument("--generated",  required=True,
                        help="Path to all_sequences.csv from generate_sequences.py")
    parser.add_argument("--benchmark",  required=True,
                        help="Path to benchmark CSV with ground truth sequences")
    parser.add_argument("--output",     default="outputs/scores.csv")
    parser.add_argument("--clone",      default=None,
                        help="Score only this clone (optional)")
    parser.add_argument("--structure",  action="store_true",
                        help="Run IgFold structure prediction (slow)")
    args = parser.parse_args()

    print(f"Loading generated sequences: {args.generated}")
    generated = load_generated(args.generated)
    print(f"Loading benchmark: {args.benchmark}")
    benchmarks = load_benchmark(args.benchmark)
    print(f"Found {len(generated)} sequence entries\n")

    # ── Group generated rows by clone ────────────────────────────────────────
    # Needed to resolve correct grafted baseline per seq_id
    from collections import defaultdict
    clone_rows: dict = defaultdict(dict)  # {clone_id: {seq_id: row}}
    for entry in generated:
        clone_id = entry.get("clone", "")
        seq_id = entry.get("seq_id", "")
        if args.clone and clone_id != args.clone:
            continue
        if clone_id in benchmarks:
            clone_rows[clone_id][seq_id] = entry

    # Baseline mapping: which seq_id provides the grafted baseline for each seq_id
    # seq_id (being scored) → seq_id of its grafted baseline sequence
    GRAFTED_BASELINE_MAP = {
        "0":  "none",    # Sapiens ran on mouse directly — no grafting baseline
        "0r": "none",
        "1":  "1",       # pipeline grafted IS its own baseline
        "2":  "1",       # Sapiens applied to seq 1
        "2r": "1",
        "3":  "3",       # lab grafted IS its own baseline
        "4":  "4",       # detected grafted IS its own baseline
        "5":  "3",       # lab back-mutations applied to seq 3
        "6":  "4",       # Sapiens applied to seq 4
        "6r": "4",
        "7":  "4",       # direct back-mutations applied to seq 4
        "8":  "8",       # stated grafted IS its own baseline
        "9":  "8",       # Sapiens applied to seq 8
        "9r": "8",
    }

    all_scores = []

    for clone_id, seq_map in clone_rows.items():
        print(f"  Scoring {clone_id}...")
        bench = benchmarks[clone_id]
        mouse_vh = bench.get("mouse_vh", "")
        mouse_vl = bench.get("mouse_vl", "")

        for seq_id, entry in seq_map.items():
            seq_label = entry.get("seq_label", "")
            vh_seq = entry.get("vh_sequence", "") or None
            vl_seq = entry.get("vl_sequence", "") or None
            vl_chain_type = entry.get("vl_chain_type", "K") or "K"

            # ── Resolve correct grafted baseline ──────────────────────────────
            baseline_id = GRAFTED_BASELINE_MAP.get(seq_id, "3")

            if baseline_id == "none":
                # seq 0/0r: no grafting happened — pass empty string so
                # compute_vernier() receives None and uses seq-0-specific logic
                grafted_vh = ""
                grafted_vl = ""
            elif baseline_id == "mouse":
                grafted_vh = mouse_vh
                grafted_vl = mouse_vl
            else:
                baseline_entry = seq_map.get(baseline_id, {})
                grafted_vh = baseline_entry.get("vh_sequence", "") or ""
                grafted_vl = baseline_entry.get("vl_sequence", "") or ""

            scores = score_one(
                clone_id=clone_id,
                seq_id=seq_id,
                seq_label=seq_label,
                vh_seq=vh_seq,
                vl_seq=vl_seq,
                vl_chain_type=vl_chain_type,
                mouse_vh=mouse_vh,
                mouse_vl=mouse_vl,
                lab_hu_vh=bench.get("hu_vh", ""),
                lab_hu_vl=bench.get("hu_vl", ""),
                lab_final_vh=bench.get("final_vh", ""),
                lab_final_vl=bench.get("final_vl", ""),
                vh_germline=entry.get("vh_germline", ""),
                vl_germline=entry.get("vl_germline", ""),
                grafted_vh=grafted_vh,
                grafted_vl=grafted_vl,
                run_structure=args.structure,
            )
            all_scores.append(scores)

    if not all_scores:
        print("No sequences scored.")
        return

    # Write output
    all_keys = []
    seen = set()
    for row in all_scores:
        for k in row:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    with open(args.output, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys, extrasaction='ignore')
        writer.writeheader()
        for row in all_scores:
            writer.writerow(row)

    print(f"\nScored {len(all_scores)} sequences → {args.output}")
    print(f"Features per sequence: {len(all_keys)}")


if __name__ == "__main__":
    main()
