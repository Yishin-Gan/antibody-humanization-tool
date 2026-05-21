"""
Build the downloadable XLSX bundle for a RunResult.

Three sheets:
  1. Glossary     — display name → scores.csv column → definition
  2. Sequences    — one row per mode/chain, residues split into FR1/CDR1/.../FR4
  3. Metrics      — every metric shown in the Feature Metrics tab, per mode
"""

from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from web import report_data as rd


# ── Region split (IMGT-based) ─────────────────────────────────────────────────


def _split_by_region(seq: str, imgt_map: list) -> dict:
    """Split a sequence into FR1/CDR1/.../FR4 by IMGT position.

    Uses the abnumber-based IMGT map already emitted by report_data._linear_to_imgt,
    so the split is consistent with the alignment block and the Vernier markers.
    """
    out = {"FR1": [], "CDR1": [], "FR2": [], "CDR2": [],
           "FR3": [], "CDR3": [], "FR4": []}
    n = min(len(seq), len(imgt_map) if imgt_map else 0)
    for i in range(n):
        pos = imgt_map[i]
        if pos is None:
            continue
        aa = seq[i]
        if   pos <= 26:  out["FR1"].append(aa)
        elif pos <= 38:  out["CDR1"].append(aa)
        elif pos <= 55:  out["FR2"].append(aa)
        elif pos <= 65:  out["CDR2"].append(aa)
        elif pos <= 104: out["FR3"].append(aa)
        elif pos <= 117: out["CDR3"].append(aa)
        else:            out["FR4"].append(aa)
    return {k: "".join(v) for k, v in out.items()}


# ── Glossary content (kept in lock-step with the Feature Metrics tab) ─────────


def _glossary_rows() -> list:
    """Return [(group, display_name, csv_column, direction, definition), ...]

    `csv_column` is the column name written by score_sequences.py to scores.csv.
    Sequences (e.g. vh_*) are listed without the chain prefix where one entry
    covers both VH/VL (the CSV uses vh_/vl_ prefixes consistently).
    """
    return [
      # ───── Identity ───────────────────────────────────────────────────────
      ("Sequence", "Mode VH sequence",          "vh_sequence",                    "—",         "Full VH amino acid sequence after this mode's processing."),
      ("Sequence", "Mode VL sequence",          "vl_sequence",                    "—",         "Full VL amino acid sequence after this mode's processing."),
      ("Sequence", "Pipeline germline VH",      "vh_germline",                    "—",         "Human germline gene selected by rank_germlines() (top-1 FR identity)."),
      ("Sequence", "Pipeline germline VL",      "vl_germline",                    "—",         "VL equivalent."),

      # ───── Humanness — OASis ──────────────────────────────────────────────
      ("Humanness — OASis", "OASis VH overall",  "oasis_vh_identity",             "higher",    "Fraction of VH 9-mer windows (FR+CDR) found in ≥10% of OAS human antibody subjects."),
      ("Humanness — OASis", "OASis FR VH",       "vh_oasis_fr_identity",          "higher",    "OASis identity using FR-only 9-mers. Primary humanness benchmark."),
      ("Humanness — OASis", "OASis CDR VH",      "vh_oasis_cdr_identity",         "higher",    "OASis identity using CDR-only 9-mers."),
      ("Humanness — OASis", "Germline FR id VH", "vh_germline_identity",          "higher",    "Fraction of VH FR IMGT positions matching the canonical germline allele."),
      ("Humanness — OASis", "OASis VL overall",  "oasis_vl_identity",             "higher",    "VL equivalent of OASis overall."),
      ("Humanness — OASis", "OASis FR VL",       "vl_oasis_fr_identity",          "higher",    "VL FR-only OASis."),
      ("Humanness — OASis", "OASis CDR VL",      "vl_oasis_cdr_identity",         "higher",    "VL CDR-only OASis."),
      ("Humanness — OASis", "Germline FR id VL", "vl_germline_identity",          "higher",    "VL FR germline identity."),

      # ───── Structure ───────────────────────────────────────────────────────
      ("Structure — ABodyBuilder2", "Conf mean VH",  "conf_mean_vh",              "higher",    "Mean ABodyBuilder2 plDDT confidence across all VH residues."),
      ("Structure — ABodyBuilder2", "Conf FR VH",    "conf_fr_mean_vh",           "higher",    "Mean confidence over FR residues only."),
      ("Structure — ABodyBuilder2", "Conf CDR VH",   "conf_cdr_mean_vh",          "higher",    "Mean confidence over CDR residues (pooled CDR1+2+3)."),
      ("Structure — ABodyBuilder2", "Conf mean VL",  "conf_mean_vl",              "higher",    "VL equivalent."),
      ("Structure — ABodyBuilder2", "Conf FR VL",    "conf_fr_mean_vl",           "higher",    "VL FR mean confidence."),
      ("Structure — ABodyBuilder2", "Conf CDR VL",   "conf_cdr_mean_vl",          "higher",    "VL CDR mean confidence."),
      ("Structure — ABodyBuilder2", "Conf min VH",   "conf_min_vh",               "higher",    "Worst single VH residue confidence — flags weak regions."),
      ("Structure — ABodyBuilder2", "Conf min VL",   "conf_min_vl",               "higher",    "Worst single VL residue confidence."),
      ("Structure — ABodyBuilder2", "Conf CDR1 VH", "conf_cdr1_mean_vh",          "higher",    "Mean confidence over VH CDR1 only."),
      ("Structure — ABodyBuilder2", "Conf CDR2 VH", "conf_cdr2_mean_vh",          "higher",    "Mean confidence over VH CDR2 only."),
      ("Structure — ABodyBuilder2", "Conf CDR3 VH", "conf_cdr3_mean_vh",          "higher",    "Mean confidence over VH CDR3 — biggest risk for long loops."),
      ("Structure — ABodyBuilder2", "Conf CDR1 VL", "conf_cdr1_mean_vl",          "higher",    "VL CDR1 mean confidence."),
      ("Structure — ABodyBuilder2", "Conf CDR2 VL", "conf_cdr2_mean_vl",          "higher",    "VL CDR2 mean confidence."),
      ("Structure — ABodyBuilder2", "Conf CDR3 VL", "conf_cdr3_mean_vl",          "higher",    "VL CDR3 mean confidence."),

      # ───── Solubility ──────────────────────────────────────────────────────
      ("Solubility — CamSol",  "CamSol VH",      "vh_camsol_score",               "higher",    "Mean CamSol intrinsic solubility score over all VH residues. >0 soluble, <0 aggregation-prone."),
      ("Solubility — CamSol",  "CamSol FR VH",   "vh_camsol_fr_score",            "higher",    "CamSol over VH FR residues."),
      ("Solubility — CamSol",  "CamSol CDR VH",  "vh_camsol_cdr_score",           "higher",    "CamSol over VH CDR residues — fixed by mouse CDR identity."),
      ("Solubility — CamSol",  "CamSol VL",      "vl_camsol_score",               "higher",    "VL equivalent."),
      ("Solubility — CamSol",  "CamSol FR VL",   "vl_camsol_fr_score",            "higher",    "CamSol over VL FR residues."),
      ("Solubility — CamSol",  "CamSol CDR VL",  "vl_camsol_cdr_score",           "higher",    "CamSol over VL CDR residues."),
      ("Solubility — CamSol",  "Hotspots VH",    "vh_camsol_hotspot_count",       "lower",     "Number of VH residues with smoothed CamSol < -0.5 (aggregation hotspots)."),
      ("Solubility — CamSol",  "Hotspots FR VH", "vh_camsol_hotspot_fr_count",    "lower",     "Hotspots restricted to VH FR."),
      ("Solubility — CamSol",  "Hotspots CDR VH","vh_camsol_hotspot_cdr_count",   "constant",  "Hotspots restricted to VH CDR — constant across modes (CDRs are unchanged)."),
      ("Solubility — CamSol",  "Hotspots VL",    "vl_camsol_hotspot_count",       "lower",     "VL total hotspots."),
      ("Solubility — CamSol",  "Hotspots FR VL", "vl_camsol_hotspot_fr_count",    "lower",     "VL FR hotspots."),
      ("Solubility — CamSol",  "Hotspots CDR VL","vl_camsol_hotspot_cdr_count",   "constant",  "VL CDR hotspots — constant."),

      # ───── Physicochemical ────────────────────────────────────────────────
      ("Physicochemical", "pI Fv",          "fv_pi",                              "—",         "Isoelectric point of the paired Fv (VH+VL)."),
      ("Physicochemical", "Net charge Fv",  "fv_net_charge_ph7",                  "—",         "Net charge at pH 7 of the paired Fv."),

      # ───── Liabilities ────────────────────────────────────────────────────
      ("FR liabilities", "Deamid FR VH",    "vh_deamidation_fr_count",            "lower",     "Number of NG/NS/NT motifs in VH FR — sites at risk of deamidation."),
      ("FR liabilities", "Deamid FR VL",    "vl_deamidation_fr_count",            "lower",     "VL FR deamidation motifs."),
      ("FR liabilities", "Oxidation FR VH", "vh_oxidation_fr_count",              "lower",     "Number of M/W residues in VH FR — sites at risk of oxidation."),
      ("FR liabilities", "Oxidation FR VL", "vl_oxidation_fr_count",              "lower",     "VL FR oxidation sites."),
      ("FR liabilities", "Isomer FR VH",    "vh_isomerization_fr_count",          "lower",     "Number of DG/DS motifs in VH FR — isomerization risk."),
      ("FR liabilities", "Isomer FR VL",    "vl_isomerization_fr_count",          "lower",     "VL FR isomerization motifs."),

      ("CDR liabilities", "Deamid CDR VH",    "vh_deamidation_cdr_count",         "constant",  "VH CDR deamidation motifs — constant across modes."),
      ("CDR liabilities", "Oxidation CDR VH", "vh_oxidation_cdr_count",           "constant",  "VH CDR oxidation sites — constant."),
      ("CDR liabilities", "Deamid CDR VL",    "vl_deamidation_cdr_count",         "constant",  "VL CDR deamidation."),
      ("CDR liabilities", "Oxidation CDR VL", "vl_oxidation_cdr_count",           "constant",  "VL CDR oxidation."),

      # ───── Vernier / back-mutations ───────────────────────────────────────
      ("Vernier / back-mutations", "Vernier mutable VH",   "vh_vernier_mutable_count",  "—",  "Count of VH Vernier IMGT positions where mouse ≠ germline (decision points)."),
      ("Vernier / back-mutations", "Vernier back-mut VH",  "vh_vernier_backmut_count",  "—",  "Vernier positions where the mouse residue was restored."),
      ("Vernier / back-mutations", "Vernier humanized VH", "vh_vernier_humanized_count","—",  "Vernier positions kept as germline residue."),
      ("Vernier / back-mutations", "Vernier mutable VL",   "vl_vernier_mutable_count",  "—",  "VL equivalent."),
      ("Vernier / back-mutations", "Vernier back-mut VL",  "vl_vernier_backmut_count",  "—",  "VL Vernier back-mutations."),
      ("Vernier / back-mutations", "Vernier humanized VL", "vl_vernier_humanized_count","—",  "VL Vernier humanizations."),
      ("Vernier / back-mutations", "Total mutable VH",     "vh_fr_mutable_count",       "—",  "All VH FR positions where mouse ≠ germline."),
      ("Vernier / back-mutations", "Total back-mut VH",    "vh_fr_backmut_count",       "—",  "All VH FR positions restored to mouse."),
      ("Vernier / back-mutations", "Total humanized VH",   "vh_fr_humanized_count",     "—",  "All VH FR positions kept as germline."),
      ("Vernier / back-mutations", "Total mutable VL",     "vl_fr_mutable_count",       "—",  "VL equivalent."),
      ("Vernier / back-mutations", "Total back-mut VL",    "vl_fr_backmut_count",       "—",  "VL FR back-mutations."),
      ("Vernier / back-mutations", "Total humanized VL",   "vl_fr_humanized_count",     "—",  "VL FR humanizations."),

      # ───── Drift (seq2 / seq6 only) ───────────────────────────────────────
      ("Sapiens drift", "Post-Sapiens germline VH",  "vh_detected_germline_post_sap_seq{sid}",          "—", "Germline gene the FR drifted toward after Sapiens (post-hoc detection)."),
      ("Sapiens drift", "Post-Sap identity VH",      "vh_detected_germline_post_sap_seq{sid}_identity", "—", "FR identity against the post-Sapiens germline."),
      ("Sapiens drift", "Drifted flag VH",           "vh_germline_drifted_seq{sid}",                    "—", "'True' if Sapiens shifted the FR toward a different gene family."),
      ("Sapiens drift", "Identity delta VH",         "vh_germline_identity_delta_seq{sid}",             "—", "Post-Sapiens minus pre-Sapiens FR identity."),
      ("Sapiens drift", "Post-Sapiens germline VL",  "vl_detected_germline_post_sap_seq{sid}",          "—", "VL equivalent."),
      ("Sapiens drift", "Post-Sap identity VL",      "vl_detected_germline_post_sap_seq{sid}_identity", "—", "VL post-Sapiens FR identity."),
      ("Sapiens drift", "Drifted flag VL",           "vl_germline_drifted_seq{sid}",                    "—", "VL drift flag."),
      ("Sapiens drift", "Identity delta VL",         "vl_germline_identity_delta_seq{sid}",             "—", "VL post − pre identity."),
    ]


# ── Metrics rows — one entry per row shown in the Feature Metrics tab ─────────

def _metrics_rows() -> list:
    """Return [(group, display_name, key_in_scores_block), ...].

    `key_in_scores_block` matches what _scores_block() emits in report_data.py.
    Special nested keys ('lia.vh_df') resolved by _deep_get below.
    """
    return [
      ("Humanness — OASis",         "OASis VH overall",  "oa_vh"),
      ("Humanness — OASis",         "OASis FR VH",       "oa_fr_vh"),
      ("Humanness — OASis",         "OASis CDR VH",      "oa_cdr_vh"),
      ("Humanness — OASis",         "Germline FR id VH", "g_vh"),
      ("Humanness — OASis",         "OASis VL overall",  "oa_vl"),
      ("Humanness — OASis",         "OASis FR VL",       "oa_fr_vl"),
      ("Humanness — OASis",         "OASis CDR VL",      "oa_cdr_vl"),
      ("Humanness — OASis",         "Germline FR id VL", "g_vl"),

      ("Structure — ABodyBuilder2", "Conf mean VH",  "cf_vh"),
      ("Structure — ABodyBuilder2", "Conf FR VH",    "cf_fr_vh"),
      ("Structure — ABodyBuilder2", "Conf CDR VH",   "cf_cdr_vh"),
      ("Structure — ABodyBuilder2", "Conf mean VL",  "cf_vl"),
      ("Structure — ABodyBuilder2", "Conf FR VL",    "cf_fr_vl"),
      ("Structure — ABodyBuilder2", "Conf CDR VL",   "cf_cdr_vl"),
      ("Structure — ABodyBuilder2", "Conf min VH",   "cf_min_vh"),
      ("Structure — ABodyBuilder2", "Conf min VL",   "cf_min_vl"),
      ("Structure — ABodyBuilder2", "Conf CDR1 VH",  "cf_cdr1_vh"),
      ("Structure — ABodyBuilder2", "Conf CDR2 VH",  "cf_cdr2_vh"),
      ("Structure — ABodyBuilder2", "Conf CDR3 VH",  "cf_cdr3_vh"),
      ("Structure — ABodyBuilder2", "Conf CDR1 VL",  "cf_cdr1_vl"),
      ("Structure — ABodyBuilder2", "Conf CDR2 VL",  "cf_cdr2_vl"),
      ("Structure — ABodyBuilder2", "Conf CDR3 VL",  "cf_cdr3_vl"),

      ("Solubility — CamSol",       "CamSol VH",       "cs_vh"),
      ("Solubility — CamSol",       "CamSol FR VH",    "cs_fr_vh"),
      ("Solubility — CamSol",       "CamSol CDR VH",   "cs_cdr_vh"),
      ("Solubility — CamSol",       "CamSol VL",       "cs_vl"),
      ("Solubility — CamSol",       "CamSol FR VL",    "cs_fr_vl"),
      ("Solubility — CamSol",       "CamSol CDR VL",   "cs_cdr_vl"),
      ("Solubility — CamSol",       "Hotspots VH",     "hs_vh"),
      ("Solubility — CamSol",       "Hotspots FR VH",  "hs_fr_vh"),
      ("Solubility — CamSol",       "Hotspots CDR VH", "hs_cdr_vh"),
      ("Solubility — CamSol",       "Hotspots VL",     "hs_vl"),
      ("Solubility — CamSol",       "Hotspots FR VL",  "hs_fr_vl"),
      ("Solubility — CamSol",       "Hotspots CDR VL", "hs_cdr_vl"),

      ("Physicochemical",           "pI Fv",          "pi"),
      ("Physicochemical",           "Net charge Fv",  "ch"),

      ("FR liabilities",            "Deamid FR VH",    "lia.vh_df"),
      ("FR liabilities",            "Deamid FR VL",    "lia.vl_df"),
      ("FR liabilities",            "Oxidation FR VH", "lia.vh_of"),
      ("FR liabilities",            "Oxidation FR VL", "lia.vl_of"),
      ("FR liabilities",            "Isomer FR VH",    "lia.vh_if"),
      ("FR liabilities",            "Isomer FR VL",    "lia.vl_if"),

      ("CDR liabilities",           "Deamid CDR VH",    "lia.vh_dc"),
      ("CDR liabilities",           "Oxidation CDR VH", "lia.vh_oc"),
      ("CDR liabilities",           "Deamid CDR VL",    "lia.vl_dc"),
      ("CDR liabilities",           "Oxidation CDR VL", "lia.vl_oc"),

      ("Vernier / back-mutations",  "Vernier mutable VH",   "vh_vern_mut"),
      ("Vernier / back-mutations",  "Vernier back-mut VH",  "vh_vern_back"),
      ("Vernier / back-mutations",  "Vernier humanized VH", "vh_vern_hum"),
      ("Vernier / back-mutations",  "Vernier mutable VL",   "vl_vern_mut"),
      ("Vernier / back-mutations",  "Vernier back-mut VL",  "vl_vern_back"),
      ("Vernier / back-mutations",  "Vernier humanized VL", "vl_vern_hum"),
      ("Vernier / back-mutations",  "Total mutable VH",     "vh_fr_mut"),
      ("Vernier / back-mutations",  "Total back-mut VH",    "vh_fr_back"),
      ("Vernier / back-mutations",  "Total humanized VH",   "vh_fr_hum"),
      ("Vernier / back-mutations",  "Total mutable VL",     "vl_fr_mut"),
      ("Vernier / back-mutations",  "Total back-mut VL",    "vl_fr_back"),
      ("Vernier / back-mutations",  "Total humanized VL",   "vl_fr_hum"),
    ]


def _deep_get(d: dict, dotted: str):
    cur = d
    for part in dotted.split("."):
        if cur is None:
            return None
        cur = cur.get(part) if isinstance(cur, dict) else None
    return cur


# ── Styling helpers ───────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="232838")
GROUP_FILL  = PatternFill("solid", fgColor="11141C")
HEADER_FONT = Font(bold=True, color="E8EAF2")
GROUP_FONT  = Font(bold=True, color="B2B8C8")
SEP_BORDER  = Border(bottom=Side(style="thin", color="6E7588"))

MODE_LABEL = {
    "pipeline":  "Pipeline",
    "preferred": "Preferred",
    "sapiens":   "Sapiens",
    "lab":       "Lab ref",
}


def _autosize(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val = cell.value
            if val is None: continue
            s = str(val)
            # cap excessively long strings (sequences) at a reasonable width
            ln = min(len(s), 60)
            if ln > max_len: max_len = ln
        ws.column_dimensions[col_letter].width = max(10, max_len + 2)


# ── Builders for each sheet ───────────────────────────────────────────────────


def _write_glossary(ws):
    ws.title = "Glossary"
    headers = ["Group", "Web display name", "CSV column", "Direction", "Definition"]
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    prev_group = None
    for grp, name, col, direction, defn in _glossary_rows():
        ws.append([grp, name, col, direction, defn])
        r = ws.max_row
        if grp != prev_group:
            for c in range(1, 6):
                ws.cell(row=r, column=c).border = SEP_BORDER
            prev_group = grp
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 80
    ws.freeze_panes = "A2"


def _write_sequences(ws, payload: dict):
    ws.title = "Sequences"
    headers = ["Mode", "Chain", "Germline", "Length", "FR1", "CDR1", "FR2", "CDR2", "FR3", "CDR3", "FR4", "Full sequence"]
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Mouse baseline first (no germline)
    mouse_vh_split = _split_by_region(payload["mouse_vh"], payload["mouse"]["vh_imgt"])
    mouse_vl_split = _split_by_region(payload["mouse_vl"], payload["mouse"]["vl_imgt"])
    ws.append(["Mouse", "VH", "", len(payload["mouse_vh"]),
               mouse_vh_split["FR1"], mouse_vh_split["CDR1"], mouse_vh_split["FR2"], mouse_vh_split["CDR2"],
               mouse_vh_split["FR3"], mouse_vh_split["CDR3"], mouse_vh_split["FR4"], payload["mouse_vh"]])
    ws.append(["Mouse", "VL", "", len(payload["mouse_vl"]),
               mouse_vl_split["FR1"], mouse_vl_split["CDR1"], mouse_vl_split["FR2"], mouse_vl_split["CDR2"],
               mouse_vl_split["FR3"], mouse_vl_split["CDR3"], mouse_vl_split["FR4"], payload["mouse_vl"]])

    for mode in payload["active_modes"]:
        d = payload["modes"][mode]
        for chain, seq_key, imgt_key, germ_key in [
            ("VH", "vh", "vh_imgt", "vh_germline"),
            ("VL", "vl", "vl_imgt", "vl_germline"),
        ]:
            seq = d[seq_key] or ""
            split = _split_by_region(seq, d[imgt_key] or [])
            ws.append([
                MODE_LABEL.get(mode, mode), chain, d[germ_key] or "", len(seq),
                split["FR1"], split["CDR1"], split["FR2"], split["CDR2"],
                split["FR3"], split["CDR3"], split["FR4"], seq,
            ])

    # Monospace font on the residue-bearing columns
    mono = Font(name="Consolas")
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=12):
        for cell in r:
            cell.font = mono
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    for col in ["E", "F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col].width = 40
    ws.column_dimensions["L"].width = 60
    ws.freeze_panes = "C2"


def _write_metrics(ws, payload: dict):
    ws.title = "Metrics"
    modes_active = list(payload["active_modes"])
    headers = ["Group", "Metric", "CSV column"] + [MODE_LABEL.get(m, m) for m in modes_active]
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Build a {display_name: csv_column} lookup from the glossary
    gloss_csv = { name: csv_col for (_g, name, csv_col, _d, _def) in _glossary_rows() }

    prev_group = None
    for group, display_name, key in _metrics_rows():
        ws.append([group, display_name, gloss_csv.get(display_name, "")] +
                  [_deep_get(payload["modes"][m]["scores"], key) for m in modes_active])
        r = ws.max_row
        if group != prev_group:
            for c in range(1, 4 + len(modes_active)):
                ws.cell(row=r, column=c).border = SEP_BORDER
            prev_group = group

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40
    for i, _ in enumerate(modes_active, start=4):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "D2"


# ── Public API ────────────────────────────────────────────────────────────────


def build_xlsx(result) -> bytes:
    payload = rd.build(result)
    wb = Workbook()
    # Sheet 1 (front): Glossary
    _write_glossary(wb.active)
    # Sheet 2: Sequences split by region
    _write_sequences(wb.create_sheet(), payload)
    # Sheet 3: All metrics
    _write_metrics(wb.create_sheet(), payload)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_tweak_xlsx(result, payload: dict, mode: str,
                     vh_edits: dict, vl_edits: dict,
                     run_structure: bool = False) -> bytes:
    """Build a tweak-focused XLSX with three sheets:
      1. Glossary      (same as full report — for column reference)
      2. Tweak Δ        (Original vs Tweaked side-by-side for one mode)
      3. Tweaked seqs   (region-split sequence rows: original + tweaked)
    """
    from web import rescore
    from web import report_data as _rd

    mode_data = payload["modes"][mode]
    chain_type = "H"   # placeholder for VH calls
    vl_chain_type = payload.get("chain_type") or "K"

    # Apply edits to both chains
    vh_orig = mode_data["vh"] or ""
    vl_orig = mode_data["vl"] or ""
    vh_seq = rescore.apply_edits(vh_orig, vh_edits)
    vl_seq = rescore.apply_edits(vl_orig, vl_edits)

    # Re-score each chain that actually changed
    vh_res = (rescore.rescore_chain(
        edited_seq=vh_seq, mouse_seq=payload["mouse_vh"], chain_type="H",
        germline_name=mode_data.get("vh_germline") or "",
        grafted_seq=mode_data.get("grafted_vh") or "")
        if vh_seq != vh_orig else None)
    vl_res = (rescore.rescore_chain(
        edited_seq=vl_seq, mouse_seq=payload["mouse_vl"], chain_type=vl_chain_type,
        germline_name=mode_data.get("vl_germline") or "",
        grafted_seq=mode_data.get("grafted_vl") or "")
        if vl_seq != vl_orig else None)

    struct = (rescore.rescore_structure(vh_seq, vl_seq, clone_id=f"tweak_{mode}")
              if run_structure and vh_seq and vl_seq else None)

    wb = Workbook()
    _write_glossary(wb.active)

    # Sheet 2: Original vs Tweaked side-by-side
    ws = wb.create_sheet("Tweak Δ")
    headers = ["Group", "Metric", "Original", "Tweaked", "Δ"]
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    orig_scores = mode_data["scores"]

    # Build a tweaked-scores dict mirroring the JS doRescore() merge
    tw_scores = dict(orig_scores)
    tw_scores["lia"] = dict(orig_scores.get("lia") or {})
    def _merge_chain(d, chain):
        if not d: return
        if chain == "VH":
            tw_scores["oa_vh"]=d["oasis"]; tw_scores["oa_fr_vh"]=d["oasis_fr"]; tw_scores["oa_cdr_vh"]=d["oasis_cdr"]
            tw_scores["g_vh"]=d["germline_id"]
            tw_scores["cs_vh"]=d["camsol"]; tw_scores["cs_fr_vh"]=d["camsol_fr"]; tw_scores["cs_cdr_vh"]=d["camsol_cdr"]
            tw_scores["hs_vh"]=(d["hs_fr"] or 0)+(d["hs_cdr"] or 0)
            tw_scores["hs_fr_vh"]=d["hs_fr"]; tw_scores["hs_cdr_vh"]=d["hs_cdr"]
            tw_scores["vh_vern_mut"]=d["vern_mut"]; tw_scores["vh_vern_back"]=d["vern_back"]; tw_scores["vh_vern_hum"]=d["vern_hum"]; tw_scores["vh_vern_other"]=d["vern_other"]
            tw_scores["vh_fr_mut"]=d["fr_mut"]; tw_scores["vh_fr_back"]=d["fr_back"]; tw_scores["vh_fr_hum"]=d["fr_hum"]; tw_scores["vh_fr_other"]=d["fr_other"]
            tw_scores["lia"].update({"vh_dc":d["lia"]["dc"],"vh_df":d["lia"]["df"],"vh_oc":d["lia"]["oc"],"vh_of":d["lia"]["of"],"vh_if":d["lia"]["if"]})
        else:
            tw_scores["oa_vl"]=d["oasis"]; tw_scores["oa_fr_vl"]=d["oasis_fr"]; tw_scores["oa_cdr_vl"]=d["oasis_cdr"]
            tw_scores["g_vl"]=d["germline_id"]
            tw_scores["cs_vl"]=d["camsol"]; tw_scores["cs_fr_vl"]=d["camsol_fr"]; tw_scores["cs_cdr_vl"]=d["camsol_cdr"]
            tw_scores["hs_vl"]=(d["hs_fr"] or 0)+(d["hs_cdr"] or 0)
            tw_scores["hs_fr_vl"]=d["hs_fr"]; tw_scores["hs_cdr_vl"]=d["hs_cdr"]
            tw_scores["vl_vern_mut"]=d["vern_mut"]; tw_scores["vl_vern_back"]=d["vern_back"]; tw_scores["vl_vern_hum"]=d["vern_hum"]; tw_scores["vl_vern_other"]=d["vern_other"]
            tw_scores["vl_fr_mut"]=d["fr_mut"]; tw_scores["vl_fr_back"]=d["fr_back"]; tw_scores["vl_fr_hum"]=d["fr_hum"]; tw_scores["vl_fr_other"]=d["fr_other"]
            tw_scores["lia"].update({"vl_dc":d["lia"]["dc"],"vl_df":d["lia"]["df"],"vl_oc":d["lia"]["oc"],"vl_of":d["lia"]["of"],"vl_if":d["lia"]["if"]})
    _merge_chain(vh_res, "VH")
    _merge_chain(vl_res, "VL")
    if struct and "_error" not in struct:
        mapping = [
            ("cf_vh","conf_mean_vh"), ("cf_vl","conf_mean_vl"),
            ("cf_fr_vh","conf_fr_mean_vh"), ("cf_fr_vl","conf_fr_mean_vl"),
            ("cf_cdr_vh","conf_cdr_mean_vh"), ("cf_cdr_vl","conf_cdr_mean_vl"),
            ("cf_cdr1_vh","conf_cdr1_mean_vh"), ("cf_cdr2_vh","conf_cdr2_mean_vh"),
            ("cf_cdr3_vh","conf_cdr3_mean_vh"),
            ("cf_cdr1_vl","conf_cdr1_mean_vl"), ("cf_cdr2_vl","conf_cdr2_mean_vl"),
            ("cf_cdr3_vl","conf_cdr3_mean_vl"),
            ("cf_min_vh","conf_min_vh"), ("cf_min_vl","conf_min_vl"),
        ]
        for short_k, long_k in mapping:
            if struct.get(long_k) is not None:
                tw_scores[short_k] = round(float(struct[long_k]), 4)

    prev_group = None
    for group, name, key in _metrics_rows():
        o = _deep_get(orig_scores, key)
        t = _deep_get(tw_scores, key)
        d = (t - o) if (o is not None and t is not None and isinstance(t, (int, float))) else None
        ws.append([group, name,
                   "" if o is None else o,
                   "" if t is None else t,
                   "" if d is None else round(d, 4)])
        r = ws.max_row
        if group != prev_group:
            for c in range(1, 6):
                ws.cell(row=r, column=c).border = SEP_BORDER
            prev_group = group

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "C2"

    # Sheet 3: Tweaked sequences split by region (original row + tweaked row)
    ws2 = wb.create_sheet("Tweaked seqs")
    headers2 = ["Variant", "Chain", "Length", "FR1", "CDR1", "FR2", "CDR2", "FR3", "CDR3", "FR4", "Full sequence"]
    ws2.append(headers2)
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    vh_imgt = mode_data["vh_imgt"]
    vl_imgt = mode_data["vl_imgt"]
    vh_imgt_t = _rd._linear_to_imgt(vh_seq, "H") if vh_res else vh_imgt
    vl_imgt_t = _rd._linear_to_imgt(vl_seq, vl_chain_type) if vl_res else vl_imgt

    for variant, vh_s, vh_i, vl_s, vl_i in [
        ("Original", vh_orig, vh_imgt, vl_orig, vl_imgt),
        ("Tweaked",  vh_seq,  vh_imgt_t, vl_seq, vl_imgt_t),
    ]:
        for label, seq, imap in [("VH", vh_s, vh_i), ("VL", vl_s, vl_i)]:
            split = _split_by_region(seq, imap or [])
            ws2.append([variant, label, len(seq),
                        split["FR1"], split["CDR1"], split["FR2"], split["CDR2"],
                        split["FR3"], split["CDR3"], split["FR4"], seq])

    mono = Font(name="Consolas")
    for r in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=4, max_col=11):
        for cell in r:
            cell.font = mono
    ws2.column_dimensions["A"].width = 11
    ws2.column_dimensions["B"].width = 6
    ws2.column_dimensions["C"].width = 8
    for col in ["D", "E", "F", "G", "H", "I", "J"]:
        ws2.column_dimensions[col].width = 36
    ws2.column_dimensions["K"].width = 60
    ws2.freeze_panes = "C2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
